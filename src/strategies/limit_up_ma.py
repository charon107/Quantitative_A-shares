import os
import pandas as pd
from openpyxl import load_workbook

# =========================
# 配置区
# =========================
ZT_CSV = "202526涨停.csv"
PARQUET_DIR = os.path.join("股价数据_parquet_fq", "kline_fq")
OUT_XLSX = "2025涨停次日开盘买入_涨停当日MA多头且不过热V2_MA20近5日单调递增_MA5近5日单调递增_多头已持续5天_未来30日最大涨幅回撤.xlsx"

NEXT_N_DAYS = 30

MA_SHORT = 7
MA_MID = 20
MA_LONG = 60

# 不过热过滤参数
LOOKBACK_DAYS = 30
MAX_RUNUP = 0.20  # 20%

# 单调递增参数
MA20_INC_DAYS = 5  # 最近5个交易日（含涨停当日）MA20严格递增
MA7_INC_DAYS = 5   # 最近5个交易日（含涨停当日）MA7严格递增

# 多头持续参数
BULL_DAYS = 5  # MA7>MA20>MA60 至少持续5天（含涨停当日）


# =========================
# 工具函数
# =========================
def to_bs_code(code6: str) -> str:
    code6 = str(code6).zfill(6)
    return "sh." + code6 if code6.startswith("6") else "sz." + code6


def load_kline(bs_code: str) -> pd.DataFrame:
    path = os.path.join(PARQUET_DIR, f"{bs_code}.parquet")
    if not os.path.exists(path):
        return pd.DataFrame()

    df = pd.read_parquet(path)
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])
    df = df.sort_values("date").reset_index(drop=True)

    for c in ["open", "high", "low", "close"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    df = df.dropna(subset=["open", "high", "low", "close"])
    return df


def add_ma(df: pd.DataFrame) -> pd.DataFrame:
    """用收盘价计算 MA7/MA20/MA60"""
    df = df.copy()
    df["MA7"] = df["close"].rolling(MA_SHORT, min_periods=MA_SHORT).mean()
    df["MA20"] = df["close"].rolling(MA_MID, min_periods=MA_MID).mean()
    df["MA60"] = df["close"].rolling(MA_LONG, min_periods=MA_LONG).mean()
    return df


def next_trade_date(kdf: pd.DataFrame, zt_date: pd.Timestamp):
    """涨停后第一个交易日"""
    idx = kdf["date"].searchsorted(zt_date + pd.Timedelta(days=1), side="left")
    if idx >= len(kdf):
        return None
    return kdf.loc[idx, "date"]


def ma_ok_on_date(kdf_ma: pd.DataFrame, d: pd.Timestamp) -> bool:
    """判断某天是否满足 MA7 > MA20 > MA60"""
    pos = kdf_ma["date"].searchsorted(d, side="left")
    if pos >= len(kdf_ma) or kdf_ma.loc[pos, "date"] != d:
        return False

    row = kdf_ma.iloc[pos]
    ma7, ma20, ma60 = row["MA7"], row["MA20"], row["MA60"]
    if pd.isna(ma7) or pd.isna(ma20) or pd.isna(ma60):
        return False
    return (ma7 > ma20) and (ma20 > ma60)


def bull_ma_continuous_n_days(kdf_ma: pd.DataFrame, d: pd.Timestamp, n: int = BULL_DAYS) -> bool:
    """
    ✅ 条件：MA7>MA20>MA60 至少已经连续 n 个交易日（含 d 当日）
    """
    pos = kdf_ma["date"].searchsorted(d, side="left")
    if pos >= len(kdf_ma) or kdf_ma.loc[pos, "date"] != d:
        return False

    start = pos - (n - 1)
    if start < 0:
        return False

    window = kdf_ma.iloc[start:pos + 1]
    if window[["MA7", "MA20", "MA60"]].isna().any().any():
        return False

    cond = (window["MA7"] > window["MA20"]) & (window["MA20"] > window["MA60"])
    return bool(cond.all())


def ma_increasing_last_n_days(kdf_ma: pd.DataFrame, d: pd.Timestamp, col: str, n: int) -> bool:
    """
    ✅ 通用：某条均线最近 n 个交易日（含 d 当日）严格单调递增
    col: "MA7"/"MA20"/"MA60"
    """
    pos = kdf_ma["date"].searchsorted(d, side="left")
    if pos >= len(kdf_ma) or kdf_ma.loc[pos, "date"] != d:
        return False

    start = pos - (n - 1)
    if start < 0:
        return False

    seq = kdf_ma.loc[start:pos, col].values
    if pd.isna(seq).any():
        return False

    return all(seq[i] < seq[i + 1] for i in range(len(seq) - 1))


def runup_ok_on_zt_date(kdf_ma: pd.DataFrame, zt_date: pd.Timestamp) -> bool:
    """
    ✅ 不过热判断（V2 口径）
    要求：涨停当日收盘价 close_t 相对以下基准的涨幅 < MAX_RUNUP
      1) 过去LOOKBACK_DAYS个交易日（不含当日）最低收盘价 min_close
      2) 过去LOOKBACK_DAYS个交易日（不含当日）MA20 的最低值 min_ma20
      base = min(min_close, min_ma20)
      runup = close_t / base - 1 < MAX_RUNUP
    """
    pos_t = kdf_ma["date"].searchsorted(zt_date, side="left")
    if pos_t >= len(kdf_ma) or kdf_ma.loc[pos_t, "date"] != zt_date:
        return False

    start = pos_t - LOOKBACK_DAYS
    end = pos_t
    if start < 0:
        return False

    close_t = kdf_ma.loc[pos_t, "close"]
    if pd.isna(close_t) or close_t <= 0:
        return False

    hist = kdf_ma.iloc[start:end].copy()
    if hist.empty:
        return False

    min_close = hist["close"].min()
    min_ma20 = hist["MA20"].min()
    if pd.isna(min_close) or pd.isna(min_ma20):
        return False

    base_price = min(min_close, min_ma20)
    if base_price <= 0:
        return False

    runup = (close_t / base_price) - 1.0
    return runup < MAX_RUNUP


def calc_nextN_metrics(kdf: pd.DataFrame, buy_date: pd.Timestamp):
    """从买入日开始取未来30交易日计算最大涨幅和最大回撤"""
    pos = kdf["date"].searchsorted(buy_date, side="left")
    if pos >= len(kdf) or kdf.loc[pos, "date"] != buy_date:
        return None, None

    window = kdf.iloc[pos: pos + NEXT_N_DAYS].copy()
    if window.empty:
        return None, None

    buy_open = window.iloc[0]["open"]
    if pd.isna(buy_open) or buy_open <= 0:
        return None, None

    max_high = window["high"].max()
    min_low = window["low"].min()

    max_return = (max_high / buy_open) - 1
    min_drawdown = (min_low / buy_open) - 1
    return float(max_return), float(min_drawdown)


def set_excel_percent_format(xlsx_path: str, percent_cols: list[str]):
    """Excel显示百分比（底层仍是数字）"""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

    for col_name in percent_cols:
        if col_name not in headers:
            continue
        col_idx = headers[col_name]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col_idx).number_format = "0.00%"

    wb.save(xlsx_path)


# =========================
# 主逻辑
# =========================
def main():
    zt = pd.read_csv(ZT_CSV, encoding="utf-8-sig")
    if "date" not in zt.columns:
        zt = pd.read_csv(ZT_CSV, encoding="utf-8")

    zt["date"] = pd.to_datetime(zt["date"], errors="coerce")
    zt = zt.dropna(subset=["date"])

    zt["code"] = zt["code"].astype(str).str.zfill(6)
    zt["name"] = zt["name"].astype(str)

    results = []

    for code6, group in zt.groupby("code", sort=False):
        bs_code = to_bs_code(code6)

        kdf = load_kline(bs_code)
        if kdf.empty:
            continue

        kdf_ma = add_ma(kdf)

        for _, row in group.iterrows():
            zt_date = row["date"]

            # 条件1：涨停当日 MA 多头
            if not ma_ok_on_date(kdf_ma, zt_date):
                continue

            # 条件2：涨停当日不过热
            if not runup_ok_on_zt_date(kdf_ma, zt_date):
                continue

            # 条件3：MA20 最近5日严格递增（含涨停当日）
            if not ma_increasing_last_n_days(kdf_ma, zt_date, col="MA20", n=MA20_INC_DAYS):
                continue

            # 条件4：MA7 最近5日严格递增（含涨停当日） ✅ 新增
            if not ma_increasing_last_n_days(kdf_ma, zt_date, col="MA7", n=MA7_INC_DAYS):
                continue

            # 条件5：MA7>MA20>MA60 至少已经连续5天（含涨停当日）
            if not bull_ma_continuous_n_days(kdf_ma, zt_date, n=BULL_DAYS):
                continue

            # 执行：次日开盘买
            buy_date = next_trade_date(kdf, zt_date)
            if buy_date is None:
                continue

            max_ret, min_dd = calc_nextN_metrics(kdf, buy_date)
            if max_ret is None:
                continue

            results.append({
                "zt_date": zt_date.strftime("%Y-%m-%d"),
                "target_date": buy_date.strftime("%Y-%m-%d"),
                "code": code6,
                "name": row["name"],
                "max_return_next30d": max_ret,
                "min_drawdown_next30d": min_dd
            })

    out = pd.DataFrame(results, columns=[
        "zt_date", "target_date", "code", "name", "max_return_next30d", "min_drawdown_next30d"
    ])

    if not out.empty:
        out["target_date_dt"] = pd.to_datetime(out["target_date"])
        out = out.sort_values(
            ["target_date_dt", "max_return_next30d"],
            ascending=[True, False]
        ).drop(columns=["target_date_dt"]).reset_index(drop=True)

    out.to_excel(OUT_XLSX, index=False)

    set_excel_percent_format(
        OUT_XLSX,
        percent_cols=["max_return_next30d", "min_drawdown_next30d"]
    )

    print(f"完成：共输出 {len(out)} 条记录 -> {OUT_XLSX}")


if __name__ == "__main__":
    main()
